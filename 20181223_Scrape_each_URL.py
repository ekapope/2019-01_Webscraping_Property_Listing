# -*- coding: utf-8 -*-
"""
Created on Sun Dec 23 20:24:30 2018

@author: Chris
"""

import urllib
from bs4 import BeautifulSoup
import pickle as pk
import openpyxl
import time
import numpy as np

#load the retrived URL list
with open('CondoLinkList.pkl','rb') as f:  # Python 3: open(..., 'rb')
    (CondoLinkList) = pk.load(f)
#check the length    
len(CondoLinkList)
###############################################################################
#Write function to retrive general info for each condo
def func_Scrape_GenInfo(QuotePage):
    print(QuotePage)
    Page = urllib.request.urlopen(QuotePage)
    Soup = BeautifulSoup(Page, 'html.parser')

    Name  = Soup.find('div', attrs={'class': 'alt'}).text
    print(Name)
    
    PriceMin = Soup.find('div', attrs={'class': 'price'}).text

# 1. Sales Information header section

    InfoMainTag = Soup.find('ul', attrs={'class': 'main-info-list'})
    InfoTagObj  = InfoMainTag.findAll('div', attrs={'class': 'value'})
    InfoCatTagObj  = InfoMainTag.findAll('div', attrs={'class': 'title'})
    
    print('---Info---')
    
    Location = ''
    Status   = ''
    Type     = ''
    Segment  = ''
    Build    = ''
    Developer= ''
    
    for k in range(0,len(InfoCatTagObj)):
        if InfoCatTagObj[k].text == 'Address':              Location = InfoTagObj[k].text
        if InfoCatTagObj[k].text == 'Status':               Status = InfoTagObj[k].text
        if InfoCatTagObj[k].text == 'Property Type':        Type = InfoTagObj[k].text
        if InfoCatTagObj[k].text == 'Segment':              Segment = InfoTagObj[k].text
        if InfoCatTagObj[k].text == 'Project Start Date':   Build = InfoTagObj[k].text
        if InfoCatTagObj[k].text == 'Developed By':         Developer = InfoTagObj[k].text
        
# 2. Basic Information section
    print('---Basic---')
    BasicMainTag = Soup.find('ul', attrs={'class': 'basic-info-list'})
    
    try:
        AreaTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item area-total'})
        Area = AreaTag.text.strip().split('\n')[1]
        print('Area = ' +  Area)
    except AttributeError :
        Area = ''
        
        
    try:        
        NTypeTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item num-unit-type'})
        NType = NTypeTag.text.strip().split('\n')[1]
        print('NType = ' + NType)
    except AttributeError :
        NType = ''

    try:          
        NUnitTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item num-unit'})
        NUnit = NUnitTag.text.strip().split('\n')[1]    
        print('NUnit = ' + NUnit)
    except AttributeError :
        NUnit = ''        

    try:    
        NFloorTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item num-floor'})
        NFloor = NFloorTag.text.strip().split('\n')[1]
        print('NFloor = ' + NFloor)
    except AttributeError :
        NFloor = ''

    try:    
        NParkingTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item num-parking'})
        NParking = NParkingTag.text.strip().split('\n')[1]
        print('NParking = ' + NParking)
    except AttributeError :
        NParking = ''        
        
    try:    
        RParkingTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item ratio-parking'})
        RParking = RParkingTag.text.strip().split('\n')[1]
        print('RParking = ' + RParking)
    except AttributeError :
        RParking = ''        

    try:    
        NLiftTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item num-lift'})
        NLift = NLiftTag.text.strip().split('\n')[1]
        print('NLift = ' + NLift)
    except AttributeError :
        NLift = ''                
        
    try:         
        PriceStartTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item price-start'})
        PriceStart = PriceStartTag.text.strip().split('\n')[1].split()[0]    
        print('PriceStart = ' + PriceStart)
    except AttributeError :
        PriceStart = ''
        
    try:        
        PriceHighTag = BasicMainTag.find('li', attrs={'class': 'basic-info-item price-end'})
        PriceHigh = PriceHighTag.text.strip().split('\n')[1].split()[0]       
        print('PriceHigh = ' + PriceHigh)
    except AttributeError :
        PriceHigh = ''
         

# 3. Image
    ImageUrl = Soup.find('div', attrs = {'class':'entity-main-image'}).find('img').get('src')
        
# 4. Location (Lat/Long)
    LocationTag = Soup.find('living-score-widget-map-canvas')
    if LocationTag == None:
        LocationTag = Soup.find('baania-map-streetview-overlay')
    
    try:
        Latitude = LocationTag.get('latitude')
    except  AttributeError :
        Latitude = ''
    
    try:
        Longitude = LocationTag.get('longitude')
    except AttributeError :
        Longitude = ''
    
# 5. Project Progress
    try:
        ProgressSect=Soup.find('div', attrs={'class': 'overall col-sm-3 col-lg-4'})
        Progress=ProgressSect.find('div', attrs={'class': 'value'}).text
    except AttributeError :
        Progress=''
        
    condoinfo=(Name, PriceMin, Location, Status,  Type, Segment, Build, Developer,
             Area, NType, NUnit,NFloor,NParking,RParking, NLift, PriceStart, PriceHigh,
             ImageUrl, Latitude, Longitude,Progress)
    
    return condoinfo,Soup
###############################################################################
#Write function to check # of rooms available inside each condo
def func_find_RoomResale(Soup):  
    # find all 'Resale' rooms in this page    
    AllRoomsTag= Soup.find('ul', attrs= {'class': 'listing-list item-list'})
    AllRooms= AllRoomsTag.findAll('li', attrs= {'class': 'listing-row item-row pie-clearfix'})
    print('------ Number of resale rooms inside this page is ' + str(len(AllRooms)))
    return AllRooms
###############################################################################
# Extract all info in each 'Resale' room
def func_Resale_info(AllRooms,i): 
    # extract data for each room
    try: Link = AllRooms[i].find('h4', attrs={'class': 'title'}).find('a').get('href') # ดึง href ของแต่ละ room
    except AttributeError : Link =''
        
    try: Title = AllRooms[i].find('h4', attrs={'class': 'title'}).text.strip().split('\n')[0]
    except AttributeError : Title =''
        
    try: Area = AllRooms[i].find('div', attrs={'class': 'usable-area'}).text.strip().split(' ', 2)[0]
    except AttributeError : Area =''
        
    try: Price = AllRooms[i].find('div', attrs={'class': 'price'}).text.strip().split('\n')[0].split()[0].replace(',', '')
    except AttributeError : Price =''
        
    try: Bed = AllRooms[i].find('div', attrs={'class': 'col column-bed'}).text.strip().split('\n')[0]
    except AttributeError : Bed =''
            
    try: Bath = AllRooms[i].find('div', attrs={'class': 'col column-bath'}).text.strip().split('\n')[0]
    except AttributeError : Bath =''
            
    try: Parking = AllRooms[i].find('div', attrs={'class': 'col column-size'}).text.strip().split('\n')[0]
    except AttributeError : Parking =''  
            
    roominfo= (Link,Title,Area,Price,Bed,Bath,Parking)
       
    return roominfo
###############################################################################
def func_find_RoomRent(Soup):  
    # find all 'Rental' rooms in this page    
    AllRoomsTag= Soup.find('ul', attrs= {'class': 'rent-list item-list'})
    AllRooms= AllRoomsTag.findAll('li', attrs= {'class': 'rent-row item-row pie-clearfix'})
    print('------ Number of rental rooms inside this page is ' + str(len(AllRooms)))
    return AllRooms
###############################################################################
# Make excel database header (first row of excel database file)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Condo_info"
dest_filename = 'Condo_info.xlsx'

current_col=0
allinfo=['Name', 'PriceMin', 'Location', 'Status','Type', 'Segment', 'Build', 'Developer',
         'CondoArea', 'NType', 'NUnit','NFloor','NParking','RParking', 'NLift', 'PriceStart', 'PriceHigh',
         'ImageUrl', 'Latitude', 'Longitude','Progress',
        'LinkS','TitleS','RoomAreaS','PriceS','BedS','BathS','ParkingS', # Resale Room info
        'LinkR','TitleR','RoomAreaR','PriceR','BedR','BathR','ParkingR'] # Rental Room info
for i in range(0,len(allinfo)):
    ws.cell(row=1,column=i+1).value = str(allinfo[i])
    current_col=current_col+1
    
wb.save(filename = dest_filename)
###############################################################################
# Get info for the page
start_time = time.time()
start_row = 2
r=start_row

for c in range(0,len(CondoLinkList)): #len(CondoLinkList):
    print('-------------- Current condo # is :' + str(c))
    condoinfo,Soup = func_Scrape_GenInfo(CondoLinkList[c])     

    # Some condos do not have Resale or Rental posting... Use try and except

    try: # check if resale room is available                
        AllRooms=func_find_RoomResale(Soup)            
        for i in range(0,len(AllRooms)): # loop all rooms in condo
            roominfo = func_Resale_info(AllRooms,i)
            for i in range(0,len(condoinfo)):
                ws.cell(row=r,column=i+1).value = condoinfo[i]              
            for i in range(0,len(roominfo)):
                ws.cell(row=r,column=len(condoinfo)+i+1).value = roominfo[i]                       
            r+=1          
        try: # check if rental room is available                
            AllRooms=func_find_RoomRent(Soup)            
            for i in range(0,len(AllRooms)): # loop all rooms in condo             
                roominfo = func_Rent_info(AllRooms,i)
                for i in range(0,len(condoinfo)):
                    ws.cell(row=r,column=i+1).value = condoinfo[i]              
                for i in range(0,len(roominfo)):
                    ws.cell(row=r,column=len(roominfo)+len(condoinfo)+i+1).value = roominfo[i]                       
                r+=1     
        except AttributeError: 
            pass       

        
    except AttributeError: # if no resale for renting
        try: # check if rental room is available                
            AllRooms=func_find_RoomRent(Soup)            
            for i in range(0,len(AllRooms)): # loop all rooms in condo                
                roominfo = func_Rent_info(AllRooms,i)
                for i in range(0,len(condoinfo)):
                    ws.cell(row=r,column=i+1).value = condoinfo[i]              
                for i in range(0,len(roominfo)):
                    ws.cell(row=r,column=len(roominfo)+len(condoinfo)+i+1).value = roominfo[i]                       
                r+=1 
                
        except AttributeError: # add only condo info            
            for i in range(0,len(condoinfo)): # add condo info for each room
                ws.cell(row=r,column=i+1).value = condoinfo[i]
            r+=1 
        
    wb.save(filename = dest_filename)         
    print('-------------- Completed Condo # '+str(c))
    print('Total time = ' + str(time.time() - start_time) + ' seconds')
    time.sleep(3)    #Add sleep time    
print('---------------- Scraping Completed ----------------')    
